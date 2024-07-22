from openpyxl import load_workbook
from campusvirtual import CampusVirtual
from actascalificacion import ActasCalificacion
import argparse, os, requests, re, sys
import re, pathlib
from selenium import webdriver
from pathlib import Path

def Chrome():
    download_dir = str(Path().absolute() / 'out')

    options = webdriver.EdgeOptions()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    options.add_experimental_option("prefs", prefs)
    options.use_chromium = True

    service = webdriver.EdgeService(executable_path=r'D:\git\uclm-eiia\py\msedgedriver.exe')
    driver = webdriver.Edge(service=service, options=options)
    return driver


def calificar_actas(notas, actas, prefijo, requisito, scale = 1., corte = 5.0, columna_total='Total del curso (Real)'):
    wb = load_workbook(filename=notas, read_only=True)
    wba = [load_workbook(filename=fname) for fname in actas]
    clean_workbooks(wba)
    for wbi in wba:
        calificacion_por_defecto(wbi)
    required_columns = calcular_columnas(wb.active, requisito)
    total_column = buscar_columna(wb.active, columna_total)
    for row in wb.active:
        givenName, sn, nota = (cell.value for cell in (row[0], row[1], row[total_column]))
        if givenName == 'Nombre':
            continue
        no_presentado = any(row[i].value == '-' for i in required_columns) if required_columns else False
        nota = 0. if nota == '-' else calificacion(nota, scale, corte)
        #print(sn, givenName, nota)
        if no_presentado and nota < 5.:
            continue
        nombre = (sn,givenName)
        if not any(poner_nota(wbi, nombre, nota) for wbi in wba):
            print('No encontrado en actas', nombre)
            

    for wbi, fname in zip(wba, actas):
        wbi.save(prefijo + fname)

def clean_workbooks(wbs):
    ''' Elimina estilos empotrados que incluyen las actas.
        Si no se hace openpyxl genera un archivo incorrecto.
    '''
    for wb in wbs:
        for row in wb.active:
            for cell in row:
                cell.style = 'Normal'

def calcular_columnas(sheet, requisito):
    header = [cell.value for cell in next(iter(sheet))]
    return [posicion_actividad(header, r) for r in requisito]

def buscar_columna(sheet, columna):
    for i, cell in enumerate(next(iter(sheet))):
        if cell.value == columna:
            return i
    return -1

def posicion_actividad(L, e):
    ra = re.compile(r'\w+:(?P<nombre>[\w\s]+)\s\(\w*\)')
    for i,t in enumerate(L):
        match = ra.match(t)
        if not match:
            continue
        if match.group('nombre') == e:
            return i


def calificacion_por_defecto(wb):
    for row in wb.active:
        if row[0].value.startswith('Número'):
            continue
        row[2].value = None
        row[3].value = 'NP'

def poner_nota(wb, nombre, nota):
    for row in wb.active:
        if igual_nombre(row[1].value, nombre):
            row[2].value = round(nota,1)
            row[2].number_format = '0.0'
            row[3].value = letra(nota)
            return True
    return False

def calificacion(nota, scale, corte):
    v = float(nota)/scale
    if v < 5. and v >= corte:
        return 5.
    return v

def letra(v):
    if v < 5.:
        return 'SS'
    if v < 7.:
        return 'AP'
    if v < 9.:
        return 'NT'
    if v < 10.:
        return 'SB'
    return 'M'

def igual_nombre(s, n):
    s = s.upper()
    a,b = n
    a, b = a.upper(), b.upper()
    if s == a + ', ' + b:
        return True
    if s == a + ' , ' + b:
        return True
    return False

try:
    from cv_cfg import USERNAME, PASSWORD 
except:
    print('''WARNING: Para recoger calificaciones necesitas un archivo cv_cfg.py con:
USERNAME='Tu nombre de usuario sin @uclm.es'
PASSWORD='Tu contraseña'
''')

if __name__ == '__main__':
    try: os.chdir(os.path.dirname(__file__))
    except: pass
    parser = argparse.ArgumentParser(description='Transfiere los datos de una hoja de CampusVirtual a las actas correspondientes')
    parser.add_argument('--ceil', nargs='?', type=float, default=5., const=4.,
                        help='aprueba (5.0) notas por encima del corte especificado')
    parser.add_argument('--noout', action='store_true',
                        help='no produce salidas de ningún tipo')
    parser.add_argument('--fetch', action='store',
                        help='descarga datos de Campus Virtual y de ActasCalificacion')
    parser.add_argument('--upload', nargs='?', default=None, const='-',
                        help='sube las actas a ActasCalificacion')
    parser.add_argument('--cv', action='store', default='cv.xlsx',
                        help='ruta de la hoja de Campus Virtual')
    parser.add_argument('--actas', nargs='+', default=['acta.xlsx'],
                        help='renderiza los datos en archivo HTML')
    parser.add_argument('--out-prefix', action='store', default='',
                        help='prefijo de las hojas de salida')
    parser.add_argument('--require', nargs='*',
                        help='tareas requeridas para considerar presentado')
    parser.add_argument('--scale', action='store', type=float, default=1.,
                        help='escalado de las notas (si en CV no son sobre 10)')
    args = parser.parse_args()
    

    if args.fetch:
        b = Chrome()
        try:
            print('La descarga/subida de archivos con Selenium es lenta, ten paciencia')
            actas = ActasCalificacion(b)
            actas.authenticate()

            print('Descargando actas vacías de', args.fetch)
            for i, f in enumerate(args.actas):
                print('Descargando acta', args.fetch, i)
                actas.download_acta(args.fetch, i)

            cv = CampusVirtual(b)
            cv.authenticate()

            print('Descargando calificaciones de', args.fetch)
            cv.calificaciones(args.fetch)
        finally:
            b.quit()

    if args.fetch:
        print('Renombrando actas descargadas')
        actas_pat = re.compile(r'CalificacionExcel( \((\d+)\))?.xlsx')

        def file_index(p):
            matched = actas_pat.search(str(p))
            suffix, index = matched.groups()
            if not suffix: return 0
            return int(index)

        actas_descargadas = pathlib.Path('download').glob('CalificacionExcel*.xlsx')
        for f, d in zip(reversed(args.actas), sorted(actas_descargadas, key=file_index, reverse=True)):
            d.rename(pathlib.Path(f))


    if args.fetch:
        print('Renombrando calificaciones descargadas')
        cv = pathlib.Path(args.cv)
        if cv.is_file(): cv.unlink()
        cv_descargado = next(pathlib.Path('download').glob(f'*{args.fetch}*.xlsx'))
        cv_descargado.rename(cv)


    if args.ceil:
        grade_ceil = True

    if not args.noout:
        print('Calculando calificaciones de', args.cv)
        calificar_actas(notas = args.cv, 
                        actas = args.actas, 
                        prefijo = args.out_prefix, 
                        requisito = args.require,
                        scale = args.scale,
                        corte = args.ceil)

    if args.upload:
        course = args.upload if args.upload != '-' else args.fetch
        if not course:
            exit()

        print("Subiendo actas de", course)
        b = Chrome()
        try:
            actas = ActasCalificacion(b)
            actas.authenticate()
            for i, fname in enumerate(args.actas):
                print("Subiendo", args.out_prefix + fname)
                actas.upload_acta(args.out_prefix + fname, course, i)
        finally:
            b.quit()
